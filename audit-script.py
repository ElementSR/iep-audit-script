import pandas as pd
import os
import glob
from datetime import datetime
from typing import Tuple
from openpyxl import load_workbook

# Final column order
FINAL_COLUMNS = [
    "Display Code", "Student Name", "Gender", "Year Level_x", "House",
    "Has Numeracy Goal", "Numeracy Goal Status", "Has Wellbeing Goal", "Wellbeing Goal Status", "Compass Meeting Count",
    "Category", "OccurredTimestamp", "Template",
    "Original Publisher", "EntryId",
    "Goal 1 - Type", "Goal 1 - Time frame", "Goal 1 - Green (goal achieved)",
    "Goal 1 - Yellow (progressing)", "Goal 1 - Red (no progress)",
    "Goal 2 - Type", "Goal 2 - Time frame", "Goal 2 - Green (goal achieved)",
    "Goal 2 - Yellow (progressing)", "Goal 2 - Red (no progress)",
    "Goal 3 - Type", "Goal 3 - Time frame", "Goal 3 - Green (goal achieved)",
    "Goal 3 - Yellow (progressing)", "Goal 3 - Red (no progress)",
    "Goal 4 - Type", "Goal 4 - Time frame", "Goal 4 - Green (goal achieved)",
    "Goal 4 - Yellow (progressing)", "Goal 4 - Red (no progress)",
    "Goal 5 - Type", "Goal 5 - Time frame", "Goal 5 - Green (goal achieved)",
    "Goal 5 - Yellow (progressing)", "Goal 5 - Red (no progress)",
    "Goal 6 - Type", "Goal 6 - Time frame", "Goal 6 - Green (goal achieved)",
    "Goal 6 - Yellow (progressing)", "Goal 6 - Red (no progress)"
]

# Helper to parse timestamps with optional seconds
TIMESTAMP_FORMATS = [
    "%d/%m/%Y %I:%M:%S %p",  # with seconds
    "%d/%m/%Y %I:%M %p"      # without seconds
]

def parse_timestamp(ts_str: str) -> datetime:
    """Try multiple timestamp formats to parse into datetime."""
    if pd.isna(ts_str) or not ts_str:
        return pd.NaT
    s = ts_str.strip()
    for fmt in TIMESTAMP_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # fallback to dateutil if needed
    try:
        return pd.to_datetime(s, dayfirst=True, errors="coerce").to_pydatetime()
    except:
        return pd.NaT


def load_latest_csv(pattern: str = "StudentChronicleOverview*.csv") -> Tuple[pd.DataFrame, str]:
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"No file found matching '{pattern}'")
    latest = max(files, key=os.path.getctime)
    return pd.read_csv(latest), latest


def parse_details_with_goal_type(details: str) -> dict:
    if pd.isna(details):
        return {}
    items = details.split("~")
    parsed = {}
    current_goal = None
    for item in items:
        parts = item.split(":", 1)
        if len(parts) == 2:
            key, value = parts[0].strip(), parts[1].strip()
            if key.startswith("Goal ") and key[5:].isdigit():
                current_goal = key
                parsed[f"{current_goal} - Type"] = value
            elif current_goal:
                parsed[f"{current_goal} - {key}"] = value
            else:
                parsed[key] = value
    return parsed


def process_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df.rename(columns={"ChronicleItemTypeTextbox": "Category", "Overview": "Template"}, inplace=True)
    # Parse timestamps robustly
    df["OccurredTimestamp"] = df["OccurredTimestamp"].astype(str).apply(parse_timestamp)
    details_df = pd.json_normalize(df.get("Details", pd.Series()).apply(parse_details_with_goal_type))
    df = df.drop(columns=["Details"], errors="ignore").merge(details_df, left_index=True, right_index=True)
    if "Original Publisher" in df.columns:
        df["Original Publisher"] = df["Original Publisher"].str.replace("Recorded by: ", "")
    return df


def add_compass_meeting_count(df: pd.DataFrame) -> pd.Series:
    return df[df["Category"] == "Compass Meetings"].groupby("Display Code").size().rename("Compass Meeting Count")


def get_goal_summary(df: pd.DataFrame, keyword: str) -> Tuple[pd.Series, pd.Series]:
    def summarize(row):
        for i in range(1, 7):
            goal_type = row.get(f"Goal {i} - Type", "")
            if keyword.lower() in str(goal_type).lower():
                for color, label in [("Green", "Green"), ("Yellow", "Yellow"), ("Red", "Red")]:
                    col = (
                        f"Goal {i} - {color} (goal achieved)" if color == "Green" else
                        f"Goal {i} - {color} (progressing)" if color == "Yellow" else
                        f"Goal {i} - {color} (no progress)"
                    )
                    if str(row.get(col)).strip().lower() == "true":
                        return True, label
                return True, None
        return False, None

    results = df.apply(summarize, axis=1)
    return results.apply(lambda x: x[0]), results.apply(lambda x: x[1])


def merge_and_filter_data(df: pd.DataFrame, compass_counts: pd.Series) -> pd.DataFrame:
    df_iep = df[df["Category"] == "Individual Education Plan (IEP)"].copy()
    df_non_iep = df[df["Category"] != "Individual Education Plan (IEP)"].copy()

    # Keep only the most recent IEP per student
    df_iep = (
        df_iep.sort_values("OccurredTimestamp", ascending=False)
              .drop_duplicates(subset="Display Code", keep="first")
              .copy()
    )

    for subset in (df_iep, df_non_iep):
        subset["Compass Meeting Count"] = subset["Display Code"].map(compass_counts).fillna(0)

    df_iep["Has Numeracy Goal"], df_iep["Numeracy Goal Status"] = get_goal_summary(df_iep, "Numeracy")
    df_iep["Has Wellbeing Goal"], df_iep["Wellbeing Goal Status"] = get_goal_summary(df_iep, "Wellbeing")

    for col in ["Has Numeracy Goal", "Numeracy Goal Status", "Has Wellbeing Goal", "Wellbeing Goal Status"]:
        df_non_iep[col] = None

    df_non_iep = (
        df_non_iep.sort_values("OccurredTimestamp", ascending=False)
                  .drop_duplicates(subset="Display Code", keep="first")
    )
    df_non_iep = df_non_iep[df_non_iep["Compass Meeting Count"] >= 4]

    combined = pd.concat([df_iep, df_non_iep], ignore_index=True)
    return combined.drop_duplicates(subset="Display Code", keep="first")


def update_master_file(df: pd.DataFrame, master_file: str, output_file: str):
    df["Display Code"] = df["Display Code"].astype(str)
    df["_ChangeFlag"] = "new"

    if os.path.exists(master_file):
        wb = load_workbook(master_file)
        ws = wb["Audit"] if "Audit" in wb.sheetnames else wb.active

        # Ensure header covers all df columns
        headers = [cell.value for cell in ws[1]]
        for col in df.columns:
            if col not in headers:
                headers.append(col)
                ws.cell(row=1, column=len(headers), value=col)
        col_idx = {col: idx+1 for idx, col in enumerate(headers)}

        # Map existing Display Code to row
        existing = {
            str(ws.cell(r, col_idx["Display Code"]).value): r
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, col_idx["Display Code"]).value is not None
        }

        for i, row in df.iterrows():
            code = row["Display Code"]
            new_ts = row["OccurredTimestamp"]
            if isinstance(new_ts, pd.Timestamp):
                new_ts = new_ts.to_pydatetime()

            if code in existing:
                r = existing[code]
                old_val = ws.cell(r, col_idx["OccurredTimestamp"]).value
                if isinstance(old_val, datetime):
                    old_ts = old_val
                elif isinstance(old_val, str):
                    old_ts = parse_timestamp(old_val)
                else:
                    old_ts = pd.NaT

                # Update if newer timestamp
                if pd.notna(new_ts) and (pd.isna(old_ts) or new_ts > old_ts):
                    df.at[i, "_ChangeFlag"] = "updated"
                    cell = ws.cell(r, col_idx["OccurredTimestamp"], value=new_ts)
                    cell.number_format = 'DD/MM/YY'
                    ws.cell(r, col_idx["Compass Meeting Count"], value=row["Compass Meeting Count"])
                    if row["Category"] == "Individual Education Plan (IEP)":
                        for col in df.columns:
                            c = col_idx[col]
                            if ws.cell(r, c).value != row[col]:
                                ws.cell(r, c, value=row[col])
                # Fallback for IEP rows with content changes
                elif row["Category"] == "Individual Education Plan (IEP)":
                    updated = False
                    for col in df.columns:
                        c = col_idx[col]
                        sheet_val = ws.cell(r, c).value
                        if sheet_val != row[col]:
                            ws.cell(r, c, value=row[col])
                            updated = True
                    df.at[i, "_ChangeFlag"] = "updated" if updated else None
                else:
                    df.at[i, "_ChangeFlag"] = None
            else:
                # Append new student
                new_r = ws.max_row + 1
                for col, val in row.items():
                    if col in col_idx:
                        ws.cell(new_r, col_idx[col], value=val)
                # _ChangeFlag remains "new"

        wb.save(master_file)
    else:
        # First run: write full sheet
        df.to_excel(master_file, sheet_name="Audit", index=False, engine="openpyxl")

    df.to_csv(output_file, index=False)


def main():
    df_raw, latest_file = load_latest_csv()
    print(f"Processing file: {latest_file}")
    suffix = datetime.today().strftime("%d%m_%H%M")
    output_file = f"Parsed_{suffix}.csv"
    master_file = "Audited_Master_IEPs.xlsx"

    df_processed = process_dataframe(df_raw)
    compass_counts = add_compass_meeting_count(df_processed)
    df_combined = merge_and_filter_data(df_processed, compass_counts)

    df_combined = df_combined.sort_values(
        by=["Compass Meeting Count", "Student Name"],
        ascending=[False, True]
    )
    df_combined = df_combined.reindex(columns=FINAL_COLUMNS, fill_value=None)

    update_master_file(df_combined, master_file, output_file)

    print(f"Saved parsed file as: {output_file}")
    print(f"Master file updated: {master_file}")


if __name__ == "__main__":
    main()
